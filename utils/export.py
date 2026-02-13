"""
Export functionality to Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from utils.helpers import format_currency, calculate_pending


def export_students_to_excel(students, filename=None):
    """Export students list to Excel"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transport_students_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Transport Students"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["ID", "Name", "Class", "Phone", "Monthly Fee", "Target Amount", 
                "Paid Amount", "Pending Amount", "Status"]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for student in students:
        pending = calculate_pending(student.target_amount, student.paid_amount)
        status = "Paid" if pending == 0 else ("Partial" if student.paid_amount > 0 else "Pending")
        
        ws.append([
            student.id,
            student.name,
            student.class_name,
            student.phone or "",
            student.monthly_fee,
            student.target_amount,
            student.paid_amount,
            pending,
            status
        ])
    
    # Adjust column widths
    column_widths = [8, 25, 15, 15, 15, 15, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save
    wb.save(filename)
    return filename


def export_payments_to_excel(payments, filename=None):
    """Export payments history to Excel"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"transport_payments_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment History"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["ID", "Student Name", "Class", "Amount", "Date", "Receipt No", "Month", "Notes"]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for payment in payments:
        ws.append([
            payment[0],  # payment id
            payment[7],  # student name
            payment[8],  # class name
            payment[2],  # amount
            payment[3],  # date
            payment[4] or "",  # receipt no
            payment[5] or "",  # month
            payment[6] or ""   # notes
        ])
    
    # Adjust column widths
    column_widths = [8, 25, 15, 12, 12, 15, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save
    wb.save(filename)
    return filename


def export_event_to_excel(event, participants, filename=None):
    """Export event details and participants to Excel"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"event_{event.name.replace(' ', '_')}_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Details"
    
    # Event info section
    ws['A1'] = "Event Name:"
    ws['B1'] = event.name
    ws['A2'] = "Description:"
    ws['B2'] = event.description or ""
    ws['A3'] = "Target Amount:"
    ws['B3'] = event.target_amount
    ws['A4'] = "Collected Amount:"
    ws['B4'] = event.collected_amount
    ws['A5'] = "Pending Amount:"
    ws['B5'] = event.target_amount - event.collected_amount
    ws['A6'] = "Deadline:"
    ws['B6'] = event.deadline or ""
    
    # Style event info
    for row in range(1, 7):
        ws[f'A{row}'].font = Font(bold=True)
    
    # Participants section
    ws['A8'] = "Participants"
    ws['A8'].font = Font(bold=True, size=14)
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["ID", "Name", "Phone", "Amount Due", "Amount Paid", "Pending", "Status"]
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style header row
    for cell in ws[10]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add participants data
    for participant in participants:
        pending = calculate_pending(participant.amount_due, participant.amount_paid)
        status = "Paid" if pending == 0 else ("Partial" if participant.amount_paid > 0 else "Pending")
        
        ws.append([
            participant.id,
            participant.name,
            participant.phone or "",
            participant.amount_due,
            participant.amount_paid,
            pending,
            status
        ])
    
    # Adjust column widths
    column_widths = [8, 25, 15, 15, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save
    wb.save(filename)
    return filename


def export_defaulters_to_excel(students, filename=None):
    """Export defaulters list to Excel"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"defaulters_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Payments"
    
    # Header styling
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["ID", "Name", "Class", "Phone", "Target Amount", "Paid Amount", 
                "Pending Amount"]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for student in students:
        pending = calculate_pending(student.target_amount, student.paid_amount)
        
        ws.append([
            student.id,
            student.name,
            student.class_name,
            student.phone or "",
            student.target_amount,
            student.paid_amount,
            pending
        ])
    
    # Adjust column widths
    column_widths = [8, 25, 15, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save
    wb.save(filename)
    return filename


def export_teacher_debt_to_excel(debts, filename=None):
    """Export teacher debt records to Excel"""
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"teacher_debt_{timestamp}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Teacher Debt"
    
    # Header styling
    header_fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = ["ID", "Teacher Name", "Amount", "Date", "Notes"]
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add data
    for debt in debts:
        ws.append([
            debt.id,
            debt.teacher_name,
            debt.amount,
            debt.debt_date,
            debt.notes or ""
        ])
    
    # Adjust column widths
    column_widths = [8, 25, 15, 15, 40]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save
    wb.save(filename)
    return filename
